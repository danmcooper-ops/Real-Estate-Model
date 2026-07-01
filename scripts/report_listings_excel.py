# scripts/report_listings_excel.py
"""Excel workbook builder for real estate listings.

Mirrors the styling conventions of scripts/report_excel.py (bold gray header,
frozen header row, per-column number formats, fixed widths). Reused by both the
Flask /api/export.xlsx route and any standalone CLI export.
"""

# (label, normalized-key, number_format, width)
_COLUMNS = [
    ('Address',       'address',       '@',            42),
    ('City',          'city',          '@',            18),
    ('State',         'state',         '@',             7),
    ('ZIP',           'zip',           '@',             8),
    ('Price',         'price',         '"$"#,##0',     14),
    ('Est. Value',    'estimate',      '"$"#,##0',     14),
    ('Beds',          'bedrooms',      '0',             7),
    ('Baths',         'bathrooms',     '0.0',           7),
    ('Sq Ft',         'squareFootage', '#,##0',        10),
    ('Lot Size',      'lotSize',       '#,##0',        10),
    ('Year Built',    'yearBuilt',     '0',            11),
    ('Type',          'propertyType',  '@',            16),
    ('Category',      'category',      '@',            13),
    ('Days on Mkt',   'daysOnMarket',  '0',            12),
    ('Listed',        'listedDate',    '@',            13),
    ('Link',          'url',           '@',            30),
]
_N_FROZEN = 1  # freeze the Address column as a row identifier


def build_listings_excel(listings, filename):
    """Write ``listings`` (normalized dicts) to a styled .xlsx at ``filename``."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Listings'

    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    header_font = Font(bold=True, color='000000', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row_header_font = Font(bold=True, color='000000')
    data_font = Font(color='000000')

    # Header row
    for ci, (label, _, _, _) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = gray_fill
        cell.alignment = header_align

    # Data rows
    for ri, listing in enumerate(listings, 2):
        for ci, (_, key, fmt, _) in enumerate(_COLUMNS, 1):
            val = listing.get(key)
            cell = ws.cell(row=ri, column=ci, value=val)
            if fmt:
                cell.number_format = fmt
            if ci <= _N_FROZEN:
                cell.fill = gray_fill
                cell.font = row_header_font
            else:
                cell.fill = white_fill
                cell.font = data_font

    # Column widths
    for ci, (_, _, _, w) in enumerate(_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'B2'
    wb.save(filename)
    return filename
